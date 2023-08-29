# -*- coding: utf-8 -*-
import sys
import ezdxf
import openpyxl

class Block():
    
    def __init__(self):
        self.block_name = None
        self.block_category = None
               
        self.groups = []        # List of Group Objects
        
    def __str__(self):
        return f'\n{self.block_name} - {self.block_category}'
        
    def get_group(self,target_group_name):
        # --- Return Group Object with name:target_group_name
        for group in self.groups:
            if (group.group_name == target_group_name):
                return group
        return None

class Group():
    
    def __init__(self):
        self.group_name = None
        
        self.ports = []     # List of Port Objects
        
    def __str__(self):
        return f'\n     {self.group_name}'    
        
class Port():
    
    def __init__(self):
        self.port_name = None       # Name of the port (str)
        self.port_position = None   # Name of the port (str)
        self.port_index = None      # Index of the port (positive int)
        
    def __str__(self):
        return f'       {self.port_name} {self.port_position} {self.port_index}'


def parse_xlsx_file(file_path):
    # --- Returns a Block Object with the Groups and Ports
    
    # --- Read Dxf
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # --- Find Groups Rows
    row_indices = []
    for row in sheet.iter_rows():
        if (row[0].value == "Group Name"):
            row_indices.append(row[0].row)
            
    row_indices.append(1000) # Max row to read
    # print(row_indices)

    # --- Extract Block Properties
    current_block = Block()
    for row in sheet.iter_rows(min_row=0, max_row=row_indices[1]-1): # Read From 0 row to row before Group0 
        if (row[0].value == 'Block Name'):
            current_block.block_name = row[1].value            
            # print(f'New Block with Name:{current_block.block_name}')
            
        if (row[0].value == 'Block Category') and (current_block is not None):
            current_block.block_category = row[1].value            
            # print(f'         Category:{current_block.block_category}') 
            
    # print(current_block)
    
    # --- Extract Groups
    for row_start,row_end in zip(row_indices[:-1],row_indices[1:]):
        new_group = Group()
        for row in sheet.iter_rows(min_row=row_start, max_row=row_end-1):
            if (row[0].value == 'Group Name'):   new_group.group_name = row[1].value                
        
        current_block.groups.append(new_group)
        # print(new_group)  
    
    
    # --- Extract Ports for every Group
    for group_index,(row_start,row_end) in enumerate(zip(row_indices[:-1],row_indices[1:])):
        current_group = current_block.groups[group_index]
        for row in sheet.iter_rows(min_row=row_start+2, max_row=row_end-1):           
            if ((row[0].value == '') or (row[0].value is None)):continue # Skip Port with Empty Name
            new_port = Port()
            new_port.port_name = row[0].value
            new_port.port_position = row[1].value
            new_port.port_index = row[2].value

            current_group.ports.append(new_port)
    
    workbook.close()        

    return current_block

def create_dxf_document(new_block,group_name,block_label,block_model,block_comment):
    # --- Get the Group Object from the Block using the group name
    selected_group = new_block.get_group(group_name)
    if (selected_group is None):
        raise ValueError(f"Group '{group_name}' not found in block '{new_block.block_name}'") 
    # print(f'Selected Group: {selected_group.group_name}')

    # --- Filter the ports by left and right and then sort them according to their index    
    ports_left = [port for port in selected_group.ports if (port.port_position == 'left') and (port.port_index is not None)]
    ports_left = sorted(ports_left, key=lambda port: port.port_index)
    ports_right = [port for port in selected_group.ports if (port.port_position == 'right') and (port.port_index is not None)]
    ports_right = sorted(ports_right, key=lambda port: port.port_index)
    
    # --- Initialize the dxf document
    doc = ezdxf.new()
    doc.header['$INSUNITS'] = 4  # Make units Millimeters(4 Corresponds to Millimeters)
    msp = doc.modelspace()
    
    # --- Create Layers
    doc.layers.new(name="Port_Text",    dxfattribs={'color': 0})         # Color 0 is Black/White
    doc.layers.new(name="block_comment", dxfattribs={'color': 4})        # Color 4 is Cyan
    doc.layers.new(name="block_label",   dxfattribs={'color': 6})        # Color 6 is Magenta
    doc.layers.new(name="block_model",   dxfattribs={'color': 8})        # Color 7 is Gray
    
    block_entities = []
    
    # --- Add Main Rectangle
    ports_dist = 10
    width = 60
    height = max([len(ports_left),len(ports_right)])*ports_dist + ports_dist  # The size of the rectangle depends on the number of ports on each size(left or right)
    half_width = width / 2
    half_height = height / 2
    lwpolyline_bounding = msp.add_lwpolyline(
        points=[
            (half_width, half_height),
            (half_width, -half_height),
            (-half_width, -half_height),
            (-half_width, half_height),
            (half_width, half_height)
        ],
    )
    block_entities.append(lwpolyline_bounding)
    
    # --- Add Ports
    ports_line_length = 10
    ports_text_xoffset = 2
    ports_text_yoffset = 1
    for port in ports_left+ports_right:
        port_name = port.port_name
        port_position = port.port_position
        port_index = port.port_index
           
        y = half_height - (port_index+1) * ports_dist  # Calculate Port y position based on the index of the port
        
        # --- Add Port Horizontal Line
        if (port_position == 'left'):
            start = (-half_width, y)
            end = (-half_width -ports_line_length, y)
        elif (port_position == 'right'):
            start = (+half_width, y)
            end = (+half_width +ports_line_length, y)        
        line_port = msp.add_line(start=start, end=end)
        block_entities.append(line_port)
        
        # --- Add Port Text
        if (port_position == 'left'):
            txt_pos = (start[0]-ports_text_xoffset,start[1]+ports_text_yoffset)     # Left side of the rectangle 
        elif (port_position == 'right'):
            txt_pos = (start[0]+ports_text_xoffset,start[1]+ports_text_yoffset)     # Right side of the rectangle
            
        mtext_port = msp.add_mtext(port_name, dxfattribs={'layer': 'Port_Text','style': "azomix"}) # azomix Font includes greek characters
        mtext_port.dxf.char_height = 3
        if (port_position == 'left'):
            mtext_port.set_location(insert=txt_pos, rotation=0, attachment_point=9)    # attachment_point=9 is MTEXT_BOTTOM_RIGHT (the anchoring of the text)
        elif (port_position == 'right'):
            mtext_port.set_location(insert=txt_pos, rotation=0, attachment_point=7)    # attachment_point=7 is MTEXT_BOTTOM_LEFT
        block_entities.append(mtext_port)
    
    # --- Add Part Label Text
    mtext_label = msp.add_mtext(block_label, dxfattribs={'layer': 'block_label','style': "azomix"})
    mtext_label.dxf.char_height = 3
    mtext_label.set_location(insert=(0,-half_height+2), rotation=0, attachment_point=8)    # attachment_point=8 is MTEXT_BOTTOM_CENTER
    block_entities.append(mtext_label)

    # --- Add Part Model Text
    mtext_model = msp.add_mtext(block_model, dxfattribs={'layer': 'block_model','style': "azomix"})
    mtext_model.dxf.char_height = 3
    mtext_model.set_location(insert=(0,+half_height-2), rotation=0, attachment_point=2)    # attachment_point=2 is MTEXT_TOP_CENTER
    block_entities.append(mtext_model)
    
    # --- Add Part Comment Text
    mtext_comment = msp.add_mtext(block_comment, dxfattribs={'layer': 'block_comment','style': "azomix"})
    mtext_comment.dxf.char_height = 5
    mtext_comment.set_location(insert=(0,0), rotation=0, attachment_point=5)    # attachment_point=5 is MTEXT_MIDDLE_CENTER
    block_entities.append(mtext_comment)
    
    # --- Initialize New Block
    new_dxf_block = doc.blocks.new(name=new_block.block_name)
    
    # --- Create DXF Block
    for entity in block_entities:
        new_entity = entity.copy()
        new_dxf_block.add_entity(new_entity)
     
    # --- Delete Shapes
    for entity in block_entities:
        msp.delete_entity(entity)
            
    # --- Add DXF Block
    msp.add_blockref(new_dxf_block.name, (0,0), dxfattribs={
        'rotation': 0,
        'xscale': 1,
        'yscale': 1,
        'zscale': 1,
    })    
       
    return doc

def main():
    # example_arguments = [
    #     'xlsx2dxf_converter.py',                      # The script name itself (position 0)
    #     'Blocks\Teco L510.xlsx',                  # file_path
    #     'power',                            # group_name
    #     'VFD23',                            # block_label
    #     'Teco L510\n210-SH1F-P 0.75kW',     # block_model    
    #     'Main\nInverter\nRollers',          # block_comment
       
    #     'output.dxf',                       # output dxf_filename
    # ]
    
    # sys.argv = example_arguments
    # print(sys.argv)
    
    file_path = sys.argv[1]
    group_name = sys.argv[2]
    block_label = sys.argv[3]
    block_model = sys.argv[4].replace("\\n", "\n")
    block_comment = sys.argv[5].replace("\\n", "\n")    
    
    dxf_filename = sys.argv[6]
    
    if (len(sys.argv) != 7):
        print("Usage: convert_dxf.py <file_path> <block_label> <block_comment> <group_name> <dxf_filename>")
        sys.exit(1)
        
    new_block = parse_xlsx_file(file_path)

    doc = create_dxf_document(new_block,group_name,block_label,block_model,block_comment)

    doc.saveas(dxf_filename)
    print(f"DXF file '{dxf_filename}' created.")


if __name__ == "__main__":
    main()











